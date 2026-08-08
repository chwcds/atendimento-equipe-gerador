# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dados completos de todos os módulos e perguntas
modules_data = [
    {
        'id': 'rotina_gerador',
        'name': 'Inspeção de Rotina Gerador',
        'short': 'Rotina Gerador',
        'desc': 'Inspeção técnica e checagem de rotina do grupo gerador'
    },
    {
        'id': 'preventiva_gerador',
        'name': 'Preventiva do Gerador',
        'short': 'Preventiva Gerador',
        'desc': 'Manutenção preventiva completa com verificação de rotina e trocas de filtros'
    },
    {
        'id': 'rotina_subestacao',
        'name': 'Inspeção de Rotina Subestação',
        'short': 'Subestação',
        'desc': 'Verificação visual, segurança e relés da subestação'
    },
    {
        'id': 'banco_capacitores',
        'name': 'Inspeção Banco de Capacitor',
        'short': 'Capacitores',
        'desc': 'Verificação das células, térmico e medidor de reativo'
    },
    {
        'id': 'eletrica',
        'name': 'Elétrica',
        'short': 'Elétrica',
        'desc': 'Inspeção dos quadros, cabos e dispositivos de proteção'
    },
    {
        'id': 'iluminacao',
        'name': 'Iluminação',
        'short': 'Iluminação',
        'desc': 'Nível de iluminância, emergência e lâmpadas do salão e áreas internas'
    }
]

questions_data = {
    'rotina_gerador': [
        {'id': 'rg_abrangencia', 'section': 'Dados técnicos', 'label': 'Esse gerador atende essa loja?', 'type': 'Opções Lado a Lado', 'options': 'Total | Parcial', 'photo': 'Sem Foto', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_fabricante', 'section': 'Dados técnicos', 'label': 'Fabricante do Gerador', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_marca_motor', 'section': 'Dados técnicos', 'label': 'Marca / Motor', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_usca_controlador', 'section': 'Dados técnicos', 'label': 'USCA / Controlador', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_potencia_kva', 'section': 'Dados técnicos', 'label': 'Potência (kVA)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_tensao_v', 'section': 'Dados técnicos', 'label': 'Tensão (V)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_qtd_baterias', 'section': 'Dados técnicos', 'label': 'Quantidade de Baterias', 'type': 'Opções Lado a Lado', 'options': '1 Bateria | 2 Baterias', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_amperagem_bat', 'section': 'Dados técnicos', 'label': 'Amperagem das baterias (Ah)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_tensao_flutuacao_bat', 'section': 'Dados técnicos', 'label': 'Qual a tensão de flutuação da bateria (V)?', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_possui_qta', 'section': 'Sistema de transferência', 'label': 'Possui QTA (Quadro de Transferência Automática)?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_qta_composicao', 'section': 'Sistema de transferência', 'label': 'Composição do QTA', 'type': 'Opções Lado a Lado', 'options': 'Contatoras | Disjuntores | Chaves', 'photo': 'Obrigatória', 'condition': 'Se Possui QTA = Sim', 'priority': '-', 'nc': '-'},
        {'id': 'rg_possui_qtm', 'section': 'Sistema de transferência', 'label': 'Gerador possui QTM?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_q1', 'section': 'Checklist de rotina', 'label': '1. Acessibilidade à sala desobstruída?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_q2', 'section': 'Checklist de rotina', 'label': '2. Existe vestígio de insetos ou roedores?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': 'Média', 'nc': 'Sim'},
        {'id': 'rg_q3', 'section': 'Checklist de rotina', 'label': '3. Controlador ligado, sem alarme e em automático?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'rg_q5', 'section': 'Checklist de rotina', 'label': '4. Gerador sem vazamento aparente?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória SÓ se a resposta for "Não"', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_q6', 'section': 'Checklist de rotina', 'label': '5. Baterias em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_q7', 'section': 'Checklist de rotina', 'label': '6. Carregador de bateria em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_q8', 'section': 'Checklist de rotina', 'label': '7. Conferido nível de água do radiador?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_q9', 'section': 'Checklist de rotina', 'label': '8. Conferido nível de óleo lubrificante?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'rg_nivel_diesel', 'section': 'Nível do diesel', 'label': '9. Nível de combustível no tanque', 'type': 'Opções Lado a Lado', 'options': '25% | 50% | 75% | 100%', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_corrente_a', 'section': 'Medições', 'label': '10. Corrente fase A (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_corrente_b', 'section': 'Medições', 'label': '11. Corrente fase B (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_corrente_c', 'section': 'Medições', 'label': '12. Corrente fase C (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'rg_usca_auto', 'section': 'Fechamento', 'label': '13. USCA na posição Auto?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'}
    ],
    'preventiva_gerador': [
        {'id': 'pg_abrangencia', 'section': 'Dados técnicos', 'label': 'Esse gerador atende essa loja?', 'type': 'Opções Lado a Lado', 'options': 'Total | Parcial', 'photo': 'Sem Foto', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_fabricante', 'section': 'Dados técnicos', 'label': 'Fabricante do Gerador', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_marca_motor', 'section': 'Dados técnicos', 'label': 'Marca / Motor', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_usca_controlador', 'section': 'Dados técnicos', 'label': 'USCA / Controlador', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_potencia_kva', 'section': 'Dados técnicos', 'label': 'Potência (kVA)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_tensao_v', 'section': 'Dados técnicos', 'label': 'Tensão (V)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_qtd_baterias', 'section': 'Dados técnicos', 'label': 'Quantidade de Baterias', 'type': 'Opções Lado a Lado', 'options': '1 Bateria | 2 Baterias', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_amperagem_bat', 'section': 'Dados técnicos', 'label': 'Amperagem das baterias (Ah)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_tensao_flutuacao_bat', 'section': 'Dados técnicos', 'label': 'Qual a tensão de flutuação da bateria (V)?', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_possui_qta', 'section': 'Sistema de transferência', 'label': 'Possui QTA (Quadro de Transferência Automática)?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_qta_composicao', 'section': 'Sistema de transferência', 'label': 'Composição do QTA', 'type': 'Opções Lado a Lado', 'options': 'Contatoras | Disjuntores | Chaves', 'photo': 'Obrigatória', 'condition': 'Se Possui QTA = Sim', 'priority': '-', 'nc': '-'},
        {'id': 'pg_possui_qtm', 'section': 'Sistema de transferência', 'label': 'Gerador possui QTM?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_q1', 'section': 'Inspeção técnica', 'label': '1. Acessibilidade à sala desobstruída?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_q2', 'section': 'Inspeção técnica', 'label': '2. Existe vestígio de insetos ou roedores?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': 'Média', 'nc': 'Sim'},
        {'id': 'pg_q3', 'section': 'Inspeção técnica', 'label': '3. Controlador ligado, sem alarme e em automático?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'pg_q5', 'section': 'Inspeção técnica', 'label': '4. Gerador sem vazamento aparente?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória SÓ se a resposta for "Não"', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_q6', 'section': 'Inspeção técnica', 'label': '5. Baterias em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_q7', 'section': 'Inspeção técnica', 'label': '6. Carregador de bateria em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_q8', 'section': 'Inspeção técnica', 'label': '7. Conferido nível de água do radiador?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_q9', 'section': 'Inspeção técnica', 'label': '8. Conferido nível de óleo lubrificante?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_nivel_diesel', 'section': 'Nível do diesel', 'label': '9. Nível de combustível no tanque', 'type': 'Opções Lado a Lado', 'options': '25% | 50% | 75% | 100%', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_filtro_oleo', 'section': 'Intervenções da Preventiva', 'label': 'Foi trocado o filtro de óleo?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_int_filtro_oleo_modelo', 'section': 'Intervenções da Preventiva', 'label': 'Qual modelo do filtro de óleo?', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_filtro_ar', 'section': 'Intervenções da Preventiva', 'label': 'Foi trocado o filtro de ar?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_int_filtro_ar_modelo', 'section': 'Intervenções da Preventiva', 'label': 'Qual modelo do filtro de ar?', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_possui_filtro_agua', 'section': 'Intervenções da Preventiva', 'label': 'Gerador possui filtro de água?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_filtro_agua_modelo', 'section': 'Intervenções da Preventiva', 'label': 'Qual modelo do filtro de água?', 'type': 'Texto', 'options': '-', 'photo': 'Obrigatória', 'condition': 'Se Gerador possui filtro de água = Sim', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_oleo_motor', 'section': 'Intervenções da Preventiva', 'label': 'Foi trocado o óleo lubrificante do motor?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'pg_int_qtd_litros_oleo', 'section': 'Intervenções da Preventiva', 'label': 'Quantos litros de óleo lubrificante foram colocados no gerador?', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_trocou_bateria', 'section': 'Intervenções da Preventiva', 'label': 'Foi trocada a bateria?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_qtd_baterias_trocadas', 'section': 'Intervenções da Preventiva', 'label': 'Quantas baterias foram trocadas?', 'type': 'Opções Lado a Lado', 'options': '1 Bateria | 2 Baterias', 'photo': 'Obrigatória', 'condition': 'Se Foi trocada a bateria = Sim', 'priority': '-', 'nc': '-'},
        {'id': 'pg_int_carregador_bat', 'section': 'Intervenções da Preventiva', 'label': 'O carregador de bateria está ligado e operando corretamente?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'pg_corrente_a', 'section': 'Medições', 'label': 'Corrente fase A (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_corrente_b', 'section': 'Medições', 'label': 'Corrente fase B (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_corrente_c', 'section': 'Medições', 'label': 'Corrente fase C (A)', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_obs_fechamento', 'section': 'Fechamento', 'label': 'Observações gerais da preventiva', 'type': 'Texto Amplo', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'pg_usca_auto', 'section': 'Fechamento', 'label': 'USCA na posição Auto?', 'type': 'Opções Lado a Lado', 'options': 'Sim | Não | Observação', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'}
    ],
    'rotina_subestacao': [
        {'id': 'sub_q1', 'section': 'Checklist da Subestação', 'label': '1. A iluminação da subestação está funcionando corretamente?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'sub_q2', 'section': 'Checklist da Subestação', 'label': '2. Os fusíveis de média tensão estão em bom estado, sem aquecimento ou oxidação?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'sub_q3', 'section': 'Checklist da Subestação', 'label': '3. A subestação possui relé de proteção?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'sub_q4_rele_operando', 'section': 'Checklist da Subestação', 'label': '4. O relé está ligado, sem falhas e operando normalmente?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': 'Se Possui relé = Sim', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'sub_q5', 'section': 'Checklist da Subestação', 'label': '5. Luvas isolantes presentes e em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'sub_q6', 'section': 'Checklist da Subestação', 'label': '6. Tapete de borracha isolante presente e em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'sub_q7', 'section': 'Checklist da Subestação', 'label': '7. Sala da subestação limpa e livre de entulhos?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Média', 'nc': 'Não'},
        {'id': 'sub_q8', 'section': 'Checklist da Subestação', 'label': '8. Existe vestígio de insetos, pássaros ou roedores?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória SÓ se a resposta for "Sim"', 'condition': '-', 'priority': 'Alta', 'nc': 'Sim'},
        {'id': 'sub_obs', 'section': 'Fechamento', 'label': 'Observações pertinentes da subestação', 'type': 'Texto Amplo', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'}
    ],
    'banco_capacitores': [
        {'id': 'cap_foto_porta_fechada', 'section': 'Visão geral', 'label': 'Foto do banco de capacitores com a porta fechada', 'type': 'Foto Exclusiva', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'cap_foto_porta_aberta', 'section': 'Visão geral', 'label': 'Foto do banco de capacitores com a porta aberta', 'type': 'Foto Exclusiva', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'cap_potencia_kvar', 'section': 'Visão geral', 'label': 'Potência total do banco em kvar', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'cap_op_ligado', 'section': 'Inspeção operacional', 'label': '1. O banco de capacitores está ligado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'cap_op_automatico', 'section': 'Inspeção operacional', 'label': '2. As células capacitivas estão em automático?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'cap_op_temperatura', 'section': 'Inspeção operacional', 'label': '3. A temperatura das células e dos cabos está normal?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'cap_op_disjuntores', 'section': 'Inspeção operacional', 'label': '4. Os disjuntores estão em bom estado e armados?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'cap_op_contatores', 'section': 'Inspeção operacional', 'label': '5. Os contatores estão operando sem ruído ou centelhamento?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'cap_op_celulas_estufamento', 'section': 'Inspeção operacional', 'label': '6. As células estão sem estufamento ou vazamento?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'cap_op_ventilador', 'section': 'Inspeção operacional', 'label': '7. O ventilador está funcionando corretamente?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Média', 'nc': 'Não'},
        {'id': 'cap_ef_reativo_medidor', 'section': 'Eficiência', 'label': 'Reativo registrado no medidor da concessionária', 'type': 'Foto Exclusiva', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'cap_obs', 'section': 'Fechamento', 'label': 'Observações pertinentes do banco de capacitores', 'type': 'Texto Amplo', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'}
    ],
    'eletrica': [
        {'id': 'el_quadros_fechados', 'section': 'Instalações elétricas', 'label': 'Quadros elétricos identificados, fechados e desobstruídos?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'el_sem_aquecimento', 'section': 'Instalações elétricas', 'label': 'Ausência de sinais de aquecimento ou cheiro de queimado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'el_cabos_bom_estado', 'section': 'Instalações elétricas', 'label': 'Cabos e conexões aparentes em bom estado?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'el_protecoes_ok', 'section': 'Instalações elétricas', 'label': 'Dispositivos de proteção sem sinais de avaria?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Crítica', 'nc': 'Não'},
        {'id': 'el_obs', 'section': 'Observações', 'label': 'Observações elétricas', 'type': 'Texto Amplo', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'}
    ],
    'iluminacao': [
        {'id': 'il_salao_vendas', 'section': 'Sistemas de iluminação', 'label': 'Iluminação do salão de vendas funcionando?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Média', 'nc': 'Não'},
        {'id': 'il_areas_internas', 'section': 'Sistemas de iluminação', 'label': 'Iluminação das áreas internas funcionando?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Média', 'nc': 'Não'},
        {'id': 'il_emergencia', 'section': 'Sistemas de iluminação', 'label': 'Iluminação de emergência em condições de uso?', 'type': 'Tri-State', 'options': 'Sim | Não | Não se aplica', 'photo': 'Obrigatória', 'condition': '-', 'priority': 'Alta', 'nc': 'Não'},
        {'id': 'il_iluminancia_lux', 'section': 'Medições', 'label': 'Iluminância média medida em lux', 'type': 'Número', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'},
        {'id': 'il_obs', 'section': 'Observações', 'label': 'Observações de iluminação', 'type': 'Texto Amplo', 'options': '-', 'photo': 'Obrigatória', 'condition': '-', 'priority': '-', 'nc': '-'}
    ]
}

wb = openpyxl.Workbook()
wb.remove(wb.active)

font_title = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
font_data = Font(name='Calibri', size=10)
font_bold = Font(name='Calibri', size=10, bold=True)

fill_primary = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
fill_header = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
fill_zebra = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
fill_card = PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid')

border_thin = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# -------------------------------------------------------------
# ABA 1: RESUMO GERAL DOS MÓDULOS
# -------------------------------------------------------------
ws_summary = wb.create_sheet(title='Resumo dos Módulos')
ws_summary.views.sheetView[0].showGridLines = True

ws_summary.merge_cells('A1:E1')
ws_summary['A1'] = 'APLICATIVO ATENDIMENTO EQUIPE GERADOR - RESUMO DE MÓDULOS'
ws_summary['A1'].font = font_title
ws_summary['A1'].fill = fill_primary
ws_summary['A1'].alignment = align_center
ws_summary.row_dimensions[1].height = 30

s_headers = ['Nº', 'Identificador (ID)', 'Nome do Módulo', 'Qtd. de Perguntas', 'Objetivo / Descrição']
for c_idx, h in enumerate(s_headers, start=1):
    cell = ws_summary.cell(row=2, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_thin
ws_summary.row_dimensions[2].height = 24

total_questions = 0
for idx, m in enumerate(modules_data, start=1):
    q_count = len(questions_data.get(m['id'], []))
    total_questions += q_count
    
    ws_summary.cell(row=idx+2, column=1, value=idx).alignment = align_center
    ws_summary.cell(row=idx+2, column=2, value=m['id']).alignment = align_left
    ws_summary.cell(row=idx+2, column=3, value=m['name']).alignment = align_left
    ws_summary.cell(row=idx+2, column=4, value=q_count).alignment = align_center
    ws_summary.cell(row=idx+2, column=5, value=m['desc']).alignment = align_left
    
    for c in range(1, 6):
        cell = ws_summary.cell(row=idx+2, column=c)
        cell.font = font_data
        cell.border = border_thin
        if idx % 2 == 0:
            cell.fill = fill_zebra
    ws_summary.row_dimensions[idx+2].height = 22

# Linha de total
tot_row = len(modules_data) + 3
ws_summary.cell(row=tot_row, column=3, value='TOTAL DE PERGUNTAS NO APLICATIVO:').alignment = Alignment(horizontal='right', vertical='center')
ws_summary.cell(row=tot_row, column=3).font = font_bold
ws_summary.cell(row=tot_row, column=4, value=total_questions).alignment = align_center
ws_summary.cell(row=tot_row, column=4).font = font_bold
ws_summary.cell(row=tot_row, column=4).fill = fill_card
ws_summary.row_dimensions[tot_row].height = 24

widths_sum = [6, 25, 32, 18, 65]
for idx, w in enumerate(widths_sum, start=1):
    ws_summary.column_dimensions[get_column_letter(idx)].width = w

# -------------------------------------------------------------
# ABA 2: TODAS AS PERGUNTAS CONSOLIDADAS
# -------------------------------------------------------------
ws_all = wb.create_sheet(title='Todas as Perguntas')
ws_all.views.sheetView[0].showGridLines = True

ws_all.merge_cells('A1:J1')
ws_all['A1'] = 'CHECKLIST DE MANUTENÇÃO - TODAS AS PERGUNTAS PARA REVISÃO'
ws_all['A1'].font = font_title
ws_all['A1'].fill = fill_primary
ws_all['A1'].alignment = align_center
ws_all.row_dimensions[1].height = 32

headers = ['#', 'Módulo', 'Seção / Etapa', 'ID Pergunta', 'Enunciado da Pergunta', 'Tipo de Entrada', 'Opções de Escolha', 'Exigência de Foto', 'Condição de Exibição', 'Prioridade']

for col_idx, h in enumerate(headers, start=1):
    cell = ws_all.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_thin
ws_all.row_dimensions[2].height = 25

row_num = 3
seq_global = 1

for m in modules_data:
    m_id = m['id']
    m_name = m['name']
    q_list = questions_data.get(m_id, [])
    
    for q in q_list:
        ws_all.cell(row=row_num, column=1, value=seq_global).alignment = align_center
        ws_all.cell(row=row_num, column=2, value=m_name).alignment = align_left
        ws_all.cell(row=row_num, column=3, value=q['section']).alignment = align_left
        ws_all.cell(row=row_num, column=4, value=q['id']).alignment = align_left
        ws_all.cell(row=row_num, column=5, value=q['label']).alignment = align_left
        ws_all.cell(row=row_num, column=6, value=q['type']).alignment = align_center
        ws_all.cell(row=row_num, column=7, value=q['options']).alignment = align_center
        ws_all.cell(row=row_num, column=8, value=q['photo']).alignment = align_center
        ws_all.cell(row=row_num, column=9, value=q['condition']).alignment = align_center
        ws_all.cell(row=row_num, column=10, value=q['priority']).alignment = align_center
        
        for c in range(1, 11):
            cell = ws_all.cell(row=row_num, column=c)
            cell.font = font_data
            cell.border = border_thin
            if row_num % 2 == 0:
                cell.fill = fill_zebra
                
        ws_all.row_dimensions[row_num].height = 24
        row_num += 1
        seq_global += 1

widths_all = [6, 25, 24, 25, 55, 20, 32, 28, 28, 14]
for idx, w in enumerate(widths_all, start=1):
    ws_all.column_dimensions[get_column_letter(idx)].width = w

# -------------------------------------------------------------
# ABAS INDIVIDUAIS POR MÓDULO
# -------------------------------------------------------------
for m in modules_data:
    m_id = m['id']
    m_name = m['name']
    m_short = m['short']
    q_list = questions_data.get(m_id, [])
    
    ws_mod = wb.create_sheet(title=m_short[:31])
    ws_mod.views.sheetView[0].showGridLines = True
    
    ws_mod.merge_cells('A1:I1')
    ws_mod['A1'] = f'CHECKLIST - {m_name.upper()} ({len(q_list)} ITENS)'
    ws_mod['A1'].font = font_title
    ws_mod['A1'].fill = fill_primary
    ws_mod['A1'].alignment = align_center
    ws_mod.row_dimensions[1].height = 30
    
    mod_headers = ['Nº', 'Seção / Etapa', 'ID Pergunta', 'Enunciado da Pergunta', 'Tipo de Entrada', 'Opções de Escolha', 'Exigência de Foto', 'Condição de Exibição', 'Prioridade']
    for col_idx, h in enumerate(mod_headers, start=1):
        cell = ws_mod.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
    ws_mod.row_dimensions[2].height = 24
    
    r_idx = 3
    for idx_q, q in enumerate(q_list, start=1):
        ws_mod.cell(row=r_idx, column=1, value=idx_q).alignment = align_center
        ws_mod.cell(row=r_idx, column=2, value=q['section']).alignment = align_left
        ws_mod.cell(row=r_idx, column=3, value=q['id']).alignment = align_left
        ws_mod.cell(row=r_idx, column=4, value=q['label']).alignment = align_left
        ws_mod.cell(row=r_idx, column=5, value=q['type']).alignment = align_center
        ws_mod.cell(row=r_idx, column=6, value=q['options']).alignment = align_center
        ws_mod.cell(row=r_idx, column=7, value=q['photo']).alignment = align_center
        ws_mod.cell(row=r_idx, column=8, value=q['condition']).alignment = align_center
        ws_mod.cell(row=r_idx, column=9, value=q['priority']).alignment = align_center
        
        for c in range(1, 10):
            cell = ws_mod.cell(row=r_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if r_idx % 2 == 0:
                cell.fill = fill_zebra
                
        ws_mod.row_dimensions[r_idx].height = 24
        r_idx += 1
        
    widths_mod = [6, 24, 25, 55, 20, 32, 28, 28, 14]
    for idx, w in enumerate(widths_mod, start=1):
        ws_mod.column_dimensions[get_column_letter(idx)].width = w

# Salvando
paths_to_save = [
    'C:/Users/Wagner/Desktop/app check list/Revisao_Perguntas_Checklist_Gerador.xlsx',
    'C:/Users/Wagner/Desktop/Revisao_Perguntas_Checklist_Gerador.xlsx',
    'C:/Users/Wagner/.gemini/antigravity/scratch/supermarket-checklist/Revisao_Perguntas_Checklist_Gerador.xlsx'
]

for p in paths_to_save:
    try:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        wb.save(p)
        print(f'Salvo com sucesso em: {p}')
    except Exception as e:
        print(f'Erro ao salvar em {p}: {e}')
